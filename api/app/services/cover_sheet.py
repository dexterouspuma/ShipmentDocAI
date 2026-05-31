"""Cover-sheet generation (PDF via ReportLab, Excel via openpyxl).

The layout here is a sensible DEFAULT placeholder. Replace `_FIELDS` and the
layout once the real cover-sheet format is provided. Generation reads from the
approved DocumentRecord's extracted fields.
"""
from __future__ import annotations

import io

from shared.models import DocumentRecord
from shared.enums import CoverSheetFormat

# (label, attribute on ExtractedDocument) — adjust to the real cover-sheet spec.
_FIELDS = [
    ("Document Type", "document_type"),
    ("Issuer", "issuer_name"),
    ("Reference No", "reference_no"),
    ("B/L Number", "bl_number"),
    ("Container No", "container_no"),
    ("Port of Loading", "port_of_loading"),
    ("Port of Discharge", "port_of_discharge"),
    ("ETA", "eta"),
    ("Gross Weight", "gross_weight"),
    ("No. of Packages", "no_of_packages"),
    ("Incoterms", "incoterms"),
    ("Freight Terms", "freight_terms"),
    ("Total Due", "total_due"),
]


def _value(rec: DocumentRecord, attr: str) -> str:
    if attr == "document_type":
        return rec.document_type.value
    field = getattr(rec.extracted, attr, None) if rec.extracted else None
    if field is None:
        return ""
    val = getattr(field, "value", field)
    return "" if val is None else str(val)


def generate(rec: DocumentRecord, fmt: CoverSheetFormat) -> tuple[bytes, str, str]:
    """Return (bytes, filename, content_type).

    Excel format produces the PIDSA freight cost-coding worksheet; PDF produces the
    simple summary cover sheet."""
    if fmt == CoverSheetFormat.PDF:
        return _pdf(rec)
    from . import freight_coding
    inputs = None
    if rec.freight_coding is not None:
        inputs = freight_coding.FreightCodingInputs(**rec.freight_coding.model_dump())
    return freight_coding.generate(rec, inputs)


def _pdf(rec: DocumentRecord) -> tuple[bytes, str, str]:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.7 * inch)
    styles = getSampleStyleSheet()
    story = [Paragraph("SHIPMENT COVER SHEET", styles["Title"]), Spacer(1, 12)]

    rows = [[label, _value(rec, attr)] for label, attr in _FIELDS]
    t = Table(rows, colWidths=[2.2 * inch, 4.3 * inch])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eef2f7")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (1, 0), (1, -1), [colors.white, colors.HexColor("#fafafa")]),
    ]))
    story.append(t)
    doc.build(story)
    return buf.getvalue(), f"cover_{rec.id}.pdf", "application/pdf"


def _excel(rec: DocumentRecord) -> tuple[bytes, str, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Cover Sheet"
    ws["A1"] = "SHIPMENT COVER SHEET"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    header_fill = PatternFill("solid", fgColor="EEF2F7")
    for label, attr in _FIELDS:
        ws.append([label, _value(rec, attr)])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=1).fill = header_fill
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    return (buf.getvalue(), f"cover_{rec.id}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
