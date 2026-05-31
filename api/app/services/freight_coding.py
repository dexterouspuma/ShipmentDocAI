"""Generate the per-shipment freight cost-coding worksheet (PIDSA format).

Reproduces the per-vendor sheet layout: header (from extracted data) + G/L coding
table (charges auto-mapped to G/L codes) + chemistry allocation block + approvals.

Allocation is a mix (per the spec):
  - auto-derived: charge totals -> charge type -> G/L code; chemistry KGS if known
  - analyst-entered: cost-center distribution + chemistry weights (left as input
    cells with live Total/% formulas)
"""
from __future__ import annotations

import io
from typing import Optional

from pydantic import BaseModel, Field

from shared.models import DocumentRecord
from . import freight_coding_config as cfg


class FreightCodingInputs(BaseModel):
    """Analyst-supplied values not present in the extracted document."""
    division: str = "MTY"                       # MTY | OEM
    gl_codes: dict[str, str] = Field(default_factory=dict)        # charge_type -> G/L code (analyst-entered)
    chemistry_kgs: dict[str, float] = Field(default_factory=dict)  # code -> KGS
    cost_center_alloc: dict[str, dict[str, float]] = Field(default_factory=dict)  # charge_type -> {cc -> amount}
    approved_by: str = ""
    manager: str = ""


def gl_code_for(key: str, inputs: "FreightCodingInputs | None" = None) -> str:
    """G/L code precedence: this document's entry -> saved default -> blank."""
    if inputs and inputs.gl_codes.get(key):
        return inputs.gl_codes[key]
    from . import gl_defaults
    saved = gl_defaults.get_defaults().get(key)
    if saved:
        return saved
    return cfg.GL_CODE_BY_CHARGE.get(key, "")


def charge_totals_for_ui(rec: DocumentRecord) -> list[dict]:
    """Charge rows with their G/L code and auto-coded total, for the review UI
    to render the cost-center allocation grid (only rows with a total > 0).

    The gl_code shown is the saved default for that charge type (analyst can
    still override per document)."""
    from . import gl_defaults
    defaults = gl_defaults.get_defaults()
    totals = _charge_totals(rec)
    rows = []
    for key, label in cfg.CHARGE_ROWS:
        rows.append({
            "charge_type": key,
            "label": label,
            "gl_code": defaults.get(key) or cfg.GL_CODE_BY_CHARGE.get(key, ""),
            "total": round(totals.get(key, 0.0), 2),
        })
    return rows


def _val(rec: DocumentRecord, attr: str) -> str:
    field = getattr(rec.extracted, attr, None) if rec.extracted else None
    if field is None:
        return ""
    v = getattr(field, "value", field)
    return "" if v is None else str(v)


def classify_charge(description: str, origin: str, mode: str) -> str:
    """Map a charge description to a charge-type key, refining Japan vs other."""
    desc = (description or "").lower()
    key = "OTHERS"
    for kw, charge_type in cfg.CHARGE_KEYWORDS:
        if kw in desc:
            key = charge_type
            break
    # Refine JAPAN variants from shipment origin.
    if "japan" in (origin or "").lower():
        if key == "AIR_FREIGHT_OTHER":
            key = "AIR_FREIGHT_JAPAN"
        elif key == "OCEAN_FREIGHT_OTHER":
            key = "OCEAN_FREIGHT_JAPAN"
    return key


def _charge_totals(rec: DocumentRecord) -> dict[str, float]:
    """Sum extracted charge amounts grouped by charge-type key."""
    totals: dict[str, float] = {k: 0.0 for k, _ in cfg.CHARGE_ROWS}
    origin = _val(rec, "port_of_loading")
    mode = rec.extracted.transport_mode.value if rec.extracted else ""
    charges = rec.extracted.charges if rec.extracted else []
    for ch in charges:
        desc = ch.description.value if ch.description else ""
        amt = ch.amount.value if ch.amount and ch.amount.value is not None else 0.0
        try:
            amt = float(amt)
        except (TypeError, ValueError):
            amt = 0.0
        key = classify_charge(desc or "", origin, mode)
        totals[key] = totals.get(key, 0.0) + amt
    return totals


def generate(rec: DocumentRecord, inputs: Optional[FreightCodingInputs] = None) -> tuple[bytes, str, str]:
    """Return (xlsx_bytes, filename, content_type)."""
    inputs = inputs or FreightCodingInputs()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Coding"

    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Header block (rows 1-7) ---
    header = [
        ("Vendor Name:", _val(rec, "issuer_name")),
        ("Vendor Invoice no.", _val(rec, "reference_no")),
        ("Awb / B/L no.", _val(rec, "bl_number") or _val(rec, "hbl_number")),
        ("Comm. Invoice no.", _val(rec, "invoice_ref")),
        ("Destination", _val(rec, "port_of_discharge")),
        ("Origin", _val(rec, "port_of_loading")),
        ("Ship Date / ETD", _val(rec, "eta") or _val(rec, "issue_date")),
    ]
    for i, (label, value) in enumerate(header, start=1):
        ws.cell(row=i, column=1, value=label).font = bold
        ws.cell(row=i, column=2, value=value)

    # --- G/L coding table ---
    start = 9
    totals = _charge_totals(rec)
    cc = cfg.COST_CENTERS
    # header row
    ws.cell(row=start, column=1, value=inputs.division).font = hdr_font
    ws.cell(row=start, column=1).fill = hdr_fill
    headers = ["G/L", "Total"] + cc + ["Total", "RATE"]
    for j, h in enumerate(headers, start=2):
        c = ws.cell(row=start, column=j, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")

    grand = 0.0
    r = start + 1
    cc_first = 4                               # column D
    cc_last = cc_first + len(cc) - 1
    for key, label in cfg.CHARGE_ROWS:
        amount = totals.get(key, 0.0)
        grand += amount
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=gl_code_for(key, inputs))
        ws.cell(row=r, column=3, value=round(amount, 2))
        # cost-center cells: filled from analyst allocation if present, else blank
        alloc = inputs.cost_center_alloc.get(key, {})
        for idx, center in enumerate(cc):
            val = alloc.get(center)
            if val is not None:
                ws.cell(row=r, column=cc_first + idx, value=val)
        first = get_column_letter(cc_first)
        last = get_column_letter(cc_last)
        ws.cell(row=r, column=cc_last + 1, value=f"=SUM({first}{r}:{last}{r})")
        for col in range(2, cc_last + 2):
            ws.cell(row=r, column=col).border = border
        r += 1
    # totals row
    ws.cell(row=r, column=1, value="Total").font = bold
    ws.cell(row=r, column=3, value=f"=SUM(C{start+1}:C{r-1})").font = bold
    for col in range(cc_first, cc_last + 2):
        L = get_column_letter(col)
        ws.cell(row=r, column=col, value=f"=SUM({L}{start+1}:{L}{r-1})").font = bold

    # --- Chemistry allocation block ---
    chem_start = r + 3
    ws.cell(row=chem_start - 1, column=1, value="Division").font = bold
    ws.cell(row=chem_start - 1, column=2, value=inputs.division)
    ws.cell(row=chem_start, column=1, value="Chemistry").font = hdr_font
    ws.cell(row=chem_start, column=1).fill = hdr_fill
    ws.cell(row=chem_start, column=2, value="KGS by Commodity / Value of Invoice").font = bold
    ws.cell(row=chem_start, column=7, value="Total").font = bold
    ws.cell(row=chem_start, column=8, value="%").font = bold

    cr = chem_start + 1
    chem_first_row = cr
    for label, code in cfg.CHEMISTRY_ROWS:
        ws.cell(row=cr, column=1, value=f"{label}  {code}")
        kgs = inputs.chemistry_kgs.get(code)
        ws.cell(row=cr, column=2, value=kgs if kgs is not None else None)   # analyst/auto
        ws.cell(row=cr, column=7, value=f"=B{cr}")
        ws.cell(row=cr, column=8, value=f"=IF($G${cr_total(len(cfg.CHEMISTRY_ROWS), chem_first_row)}=0,0,G{cr}/$G${cr_total(len(cfg.CHEMISTRY_ROWS), chem_first_row)})")
        cr += 1
    total_row = cr
    ws.cell(row=total_row, column=1, value="Total").font = bold
    ws.cell(row=total_row, column=7, value=f"=SUM(G{chem_first_row}:G{total_row-1})").font = bold
    ws.cell(row=total_row, column=8, value=f"=IF(G{total_row}=0,0,SUM(H{chem_first_row}:H{total_row-1}))").font = bold

    # --- Approval block ---
    appr = total_row + 3
    ws.cell(row=appr, column=1, value="Approved by:").font = bold
    ws.cell(row=appr, column=2, value=inputs.approved_by)
    ws.cell(row=appr + 1, column=1, value="Sign / Date:").font = bold
    ws.cell(row=appr + 2, column=1, value="Approved by Manager:").font = bold
    ws.cell(row=appr + 2, column=2, value=inputs.manager)
    ws.cell(row=appr + 3, column=1, value="Sign / Date:").font = bold

    # widths
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    for col in range(3, cc_last + 3):
        ws.column_dimensions[get_column_letter(col)].width = 11

    buf = io.BytesIO()
    wb.save(buf)
    return (buf.getvalue(), f"freight_coding_{rec.id}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def cr_total(num_rows: int, first_row: int) -> int:
    """Row number of the chemistry Total row (first_row + num_rows)."""
    return first_row + num_rows
