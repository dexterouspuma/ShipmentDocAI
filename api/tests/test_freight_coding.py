"""Tests for the PIDSA freight cost-coding worksheet generator."""
import io

from openpyxl import load_workbook

from shared.models import DocumentRecord, ExtractedDocument, ExtractedField, ChargeLine
from shared.enums import DocumentType, TransportMode
from app.services import freight_coding
from app.services.freight_coding import FreightCodingInputs, classify_charge


def _f(v):
    return ExtractedField(value=v, confidence=0.95)


def _carrier_invoice_record(origin="Shanghai, CN") -> DocumentRecord:
    ed = ExtractedDocument(document_type=DocumentType.CARRIER_INVOICE,
                           transport_mode=TransportMode.OCEAN_FCL)
    ed.issuer_name = _f("OCEAN NETWORK EXPRESS")
    ed.reference_no = _f("FRT-448219")
    ed.bl_number = _f("ONEY140604281955")
    ed.invoice_ref = _f("CI-48213")
    ed.port_of_loading = _f(origin)
    ed.port_of_discharge = _f("Los Angeles, US")
    ed.charges = [
        ChargeLine(description=_f("Ocean Freight"), amount=_f(1850.0)),
        ChargeLine(description=_f("Terminal Handling Charge (THC)"), amount=_f(300.0)),
        ChargeLine(description=_f("Customs Clearance"), amount=_f(150.0)),
        ChargeLine(description=_f("Documentation Fee"), amount=_f(80.0)),
    ]
    return DocumentRecord(original_filename="ci.pdf", blob_path_raw="local://ci.pdf",
                          document_type=DocumentType.CARRIER_INVOICE, extracted=ed)


def test_classify_charge_japan_vs_other():
    assert classify_charge("Ocean Freight", "Shanghai, CN", "ocean_fcl") == "OCEAN_FREIGHT_OTHER"
    assert classify_charge("Ocean Freight", "KOBE, JAPAN", "ocean_fcl") == "OCEAN_FREIGHT_JAPAN"
    assert classify_charge("Air Freight", "NARITA, JAPAN", "air") == "AIR_FREIGHT_JAPAN"
    assert classify_charge("Customs Duty", "x", "x") == "CUSTOMS_DUTY"
    assert classify_charge("Mystery line", "x", "x") == "OTHERS"


def test_charge_totals_grouping():
    rec = _carrier_invoice_record()
    totals = freight_coding._charge_totals(rec)
    assert totals["OCEAN_FREIGHT_OTHER"] == 1850.0
    assert totals["CUSTOM_BROKER_SERVICES"] == 150.0
    # THC + Documentation both classify to FREIGHT_LOCAL
    assert totals["FREIGHT_LOCAL"] == 380.0


def test_generate_writes_cost_center_and_chemistry_inputs():
    rec = _carrier_invoice_record()
    inputs = FreightCodingInputs(
        division="OEM",
        cost_center_alloc={"OCEAN_FREIGHT_OTHER": {"13426": 1000.0, "13483": 850.0}},
        chemistry_kgs={"M3521": 2368.0},
        approved_by="A. Analyst", manager="M. Boss",
    )
    data, _, _ = freight_coding.generate(rec, inputs)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Coding"]
    # find the ocean freight row and confirm the allocated cost-center cells
    ocean_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "OCEAN FREIGHT (other than JAPAN)":
            ocean_row = r
            break
    assert ocean_row is not None
    assert ws.cell(row=ocean_row, column=4).value == 1000.0   # 13426 -> col D
    assert ws.cell(row=ocean_row, column=5).value == 850.0    # 13483 -> col E
    # chemistry KGS for LI-ION CYLINDER (M3521) written into the chem block
    assert any(c.value == 2368.0 for row in ws.iter_rows() for c in row)


def test_charge_totals_for_ui_shape():
    rows = freight_coding.charge_totals_for_ui(_carrier_invoice_record())
    by_type = {r["charge_type"]: r for r in rows}
    assert by_type["OCEAN_FREIGHT_OTHER"]["total"] == 1850.0
    # G/L default is blank until the authoritative mapping is provided.
    assert by_type["OCEAN_FREIGHT_OTHER"]["gl_code"] == ""


def test_generate_valid_xlsx_with_header_and_gl_codes():
    rec = _carrier_invoice_record()
    data, filename, content_type = freight_coding.generate(
        rec, FreightCodingInputs(division="MTY", approved_by="A. Analyst"))
    assert filename.endswith(".xlsx")
    assert data[:2] == b"PK"

    wb = load_workbook(io.BytesIO(data))
    ws = wb["Coding"]
    # Header pulled from extracted data
    assert ws["A1"].value == "Vendor Name:"
    assert ws["B1"].value == "OCEAN NETWORK EXPRESS"
    assert ws["B3"].value == "ONEY140604281955"      # AWB/BL
    assert ws["B5"].value == "Los Angeles, US"        # destination
    # G/L coding table: ocean freight row carries the total; G/L blank by default.
    found_ocean = False
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "OCEAN FREIGHT (other than JAPAN)":
            assert row[1] in (None, "")              # G/L blank until analyst sets it
            assert row[2] == 1850.0                  # total amount
            found_ocean = True
    assert found_ocean


def test_analyst_gl_code_override_is_written():
    rec = _carrier_invoice_record()
    inputs = FreightCodingInputs(gl_codes={"OCEAN_FREIGHT_OTHER": "9001234567"})
    data, _, _ = freight_coding.generate(rec, inputs)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Coding"]
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "OCEAN FREIGHT (other than JAPAN)":
            assert row[1] == "9001234567"            # analyst-entered code persists
            return
    raise AssertionError("ocean freight row not found")
