"""Tests for persisted G/L code defaults + precedence."""
import io

from openpyxl import load_workbook

from shared.models import DocumentRecord, ExtractedDocument, ExtractedField, ChargeLine
from shared.enums import DocumentType, TransportMode
from app.services import freight_coding, gl_defaults
from app.services.gl_defaults import merge_gl_codes


def test_merge_precedence_doc_over_defaults():
    defaults = {"OCEAN_FREIGHT_OTHER": "DEF-1", "FREIGHT_LOCAL": "DEF-2"}
    doc = {"OCEAN_FREIGHT_OTHER": "DOC-1"}        # doc overrides one
    merged = merge_gl_codes(doc, defaults)
    assert merged["OCEAN_FREIGHT_OTHER"] == "DOC-1"   # document wins
    assert merged["FREIGHT_LOCAL"] == "DEF-2"          # default fills the rest


def test_save_and_get_defaults_roundtrip(tmp_path, monkeypatch):
    # Redirect the defaults file into a temp dir so we don't touch real state.
    monkeypatch.setattr(gl_defaults, "_path", lambda: tmp_path / "gl_defaults.json")
    gl_defaults.save_defaults({"OCEAN_FREIGHT_OTHER": "5313072008", "OTHERS": ""})
    saved = gl_defaults.get_defaults()
    assert saved == {"OCEAN_FREIGHT_OTHER": "5313072008"}   # blank dropped


def test_defaults_applied_in_generated_sheet(tmp_path, monkeypatch):
    monkeypatch.setattr(gl_defaults, "_path", lambda: tmp_path / "gl_defaults.json")
    gl_defaults.save_defaults({"OCEAN_FREIGHT_OTHER": "DEFAULT-OCEAN"})

    ed = ExtractedDocument(document_type=DocumentType.CARRIER_INVOICE,
                           transport_mode=TransportMode.OCEAN_FCL)
    ed.port_of_loading = ExtractedField(value="Shanghai, CN", confidence=0.9)
    ed.charges = [ChargeLine(description=ExtractedField(value="Ocean Freight", confidence=0.9),
                             amount=ExtractedField(value=1850.0, confidence=0.9))]
    rec = DocumentRecord(original_filename="ci.pdf", blob_path_raw="local://ci.pdf",
                         document_type=DocumentType.CARRIER_INVOICE, extracted=ed)

    # No per-document override -> the saved default should appear.
    data, _, _ = freight_coding.generate(rec)
    ws = load_workbook(io.BytesIO(data))["Coding"]
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "OCEAN FREIGHT (other than JAPAN)":
            assert row[1] == "DEFAULT-OCEAN"
            return
    raise AssertionError("ocean freight row not found")
